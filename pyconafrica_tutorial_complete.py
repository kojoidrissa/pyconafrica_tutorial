import json
import openpyxl
from openpyxl import Workbook

# get workbook object; the data_only option captures the current value of any formulae
# Your file should be in the same directory as this code, OR you need to provide the FULL path to that file, not just it's name
wb = openpyxl.load_workbook('pyconafrica_tutorial.xlsx', data_only= True)

# WB -> WS -> Cell
demo_worksheet = wb.get_sheet_by_name("clean_data")

# Step 3: Set Comprehension
employee_ids = {
    row[0].value
    for row in demo_worksheet.rows # returns generator object
    if row[0].value != 'employee_num'
}

# Step 3 Creating a set with a for loop
# employee_ids = set()
# for row in demo_worksheet.rows: # returns a generator object
#     if row[0].value != 'employee_num':
#         employee_ids.add(row[0].value) # pulling data from a cell

# Step 4
# Create a dictionary of each employee and their info for the month
# This is the slowest step

employee_aggregate = {}
for employee in employee_ids:
    # list comprehension: I'll want to sum the hours later
    hours = [
        row[6].value
        for row in demo_worksheet.rows
        if employee == row[0].value
    ]

    # set comprehension: each employee should belong to only ONE cost center
    cost_center = {row[1].value for row in demo_worksheet.rows if employee == row[0].value}
    division = {str(row[2].value) for row in demo_worksheet.rows if employee == row[0].value}
    manager = {row[3].value for row in demo_worksheet.rows if employee == row[0].value}
    
    assert len(cost_center) == len(division) == len(manager) == 1
    
    employee_aggregate[employee]={
        "hours": sum(hours),
        "cost_center": list(cost_center)[0],
        "division": list(division)[0],
        "manager": list(manager)[0]
    }

# Step 5
# Create output workbook, then output worksheet
output_book = Workbook() #OpenPyXL object
output_sheet = output_book.create_sheet("Aggregate Time",0)

# Building the Output Header: Specific Cell References
header = [
    demo_worksheet["A1"].value,
    demo_worksheet["B1"].value,
    demo_worksheet["C1"].value,
    demo_worksheet["D1"].value,
    demo_worksheet["G1"].value
]

# Create output data construct & append header 
output_data = []
output_data.append(header)

# Building new rows, then appending them to output data construct
for employee in employee_aggregate: #iterating over dictionary keys
    new_row = []
    new_row.append(employee)
    new_row.append(employee_aggregate[employee]['cost_center'])
    new_row.append(employee_aggregate[employee]['division'])
    new_row.append(employee_aggregate[employee]['manager'])
    new_row.append(employee_aggregate[employee]['hours'])
    output_data.append(new_row)

# Step 6
# Write data to sheet object
for row in output_data:
    rowIn = output_data.index(row)
    for col in range(len(output_data[0])):
        # Adding 1 because spreadsheets count from 1, not 0
        output_sheet.cell(row = rowIn+1, column = col+1).value = output_data[rowIn][col]

# You don't have an actual spreadsheet until you do this
output_book.save (filename = "pyconafrica_tutorial_done.xlsx")
print("File created: pyconafrica_tutorial_done.xlsx ")

# Step 7
with open("aggregate_time.json", 'w') as f:
    json.dump(employee_aggregate, f, sort_keys=True, indent=4)
print("File created: aggregate_time.json")